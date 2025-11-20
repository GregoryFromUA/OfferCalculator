import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from offer_calculator import OfferCalculator, OfferResources, BaseRates, load_existing_offers
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import subprocess
import json
from io import BytesIO


st.set_page_config(page_title="Offer Calculator", page_icon="💎", layout="wide")

st.title("💎 Offer Calculator")
st.markdown("### Расчёт параметров офферов для мобильной игры")

if 'offers' not in st.session_state:
    st.session_state.offers = []

if 'base_rates' not in st.session_state:
    st.session_state.base_rates = BaseRates()

tabs = st.tabs(["📊 Новый оффер", "📋 Существующие офферы", "📈 Сравнение", "⚙️ Настройки", "💾 Экспорт"])

# ============== TAB 1: Новый оффер ==============
with tabs[0]:
    st.header("Создание нового оффера")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Основная информация")
        offer_name = st.text_input("Название оффера", value="New Offer")
        pack_price = st.number_input("Цена пака ($)", min_value=0.0, value=1.99, step=0.01)

        st.subheader("Ресурсы")
        gems = st.number_input("GEMS", min_value=0, value=0, step=10)
        skip = st.number_input("SKIP", min_value=0, value=0, step=5)
        tnt = st.number_input("TNT", min_value=0, value=0, step=10)
        nitro = st.number_input("NITRO", min_value=0, value=0, step=10)
        no_ads = st.number_input("NO ADS (1 or 0)", min_value=0, max_value=1, value=0, step=1)

        st.subheader("Сундуки")
        chest_type = st.selectbox("Тип сундука", ["NoChest", "Small", "Big"])
        chest_amount = st.number_input("Количество сундуков", min_value=0, value=0, step=1)

    with col2:
        st.subheader("📊 Расчёт оффера")

        if st.button("🔢 Рассчитать", type="primary"):
            calculator = OfferCalculator(st.session_state.base_rates)
            resources = OfferResources(
                gems=gems,
                skip=skip,
                tnt=tnt,
                nitro=nitro,
                no_ads=no_ads,
                chest_type=chest_type,
                chest_amount=chest_amount
            )

            offer_data = calculator.calculate_offer(offer_name, resources, pack_price)

            st.success("✅ Оффер рассчитан!")

            metric_col1, metric_col2, metric_col3 = st.columns(3)

            with metric_col1:
                st.metric("Base Value", f"${offer_data['Base Value ($)']}")
                st.metric("Pack Price", f"${offer_data['Pack Price ($)']}")

            with metric_col2:
                st.metric("Discount", f"{offer_data['Discount (%)']}%")
                st.metric("ROI", f"{offer_data['ROI (%)']}%")

            with metric_col3:
                st.metric("Multiplier", offer_data['Multiplier'])
                st.metric("Pack Type", offer_data['Pack Type'])

            st.info(f"**Value Badge:** {offer_data['Value Badge']}")
            st.info(f"**Min SkipIts:** {offer_data['Min SkipIts']}")

            st.divider()

            st.subheader("📝 Полная информация")
            df = pd.DataFrame([offer_data])
            st.dataframe(df, use_container_width=True)

            if st.button("💾 Добавить в список офферов"):
                st.session_state.offers.append(offer_data)
                st.success(f"Оффер '{offer_name}' добавлен!")

# ============== TAB 2: Существующие офферы ==============
with tabs[1]:
    st.header("Существующие офферы")

    if st.button("📂 Загрузить из OfferCalculator_1.xlsx"):
        try:
            df = load_existing_offers("OfferCalculator_1.xlsx")
            st.success(f"Загружено {len(df)} офферов!")
            st.dataframe(df, use_container_width=True)
        except Exception as e:
            st.error(f"Ошибка загрузки: {e}")

    if st.session_state.offers:
        st.subheader(f"Список офферов ({len(st.session_state.offers)})")
        df_offers = pd.DataFrame(st.session_state.offers)
        st.dataframe(df_offers, use_container_width=True)

        if st.button("🗑️ Очистить список"):
            st.session_state.offers = []
            st.rerun()
    else:
        st.info("Список офферов пуст. Создайте оффер в разделе 'Новый оффер'")

# ============== TAB 3: Сравнение ==============
with tabs[2]:
    st.header("Сравнение офферов")

    if st.session_state.offers:
        df_offers = pd.DataFrame(st.session_state.offers)

        comparison_col1, comparison_col2 = st.columns(2)

        with comparison_col1:
            st.subheader("📊 ROI по офферам")
            fig_roi = px.bar(
                df_offers,
                x="Name",
                y="ROI (%)",
                color="ROI (%)",
                color_continuous_scale="Viridis",
                title="ROI по офферам"
            )
            st.plotly_chart(fig_roi, use_container_width=True)

        with comparison_col2:
            st.subheader("💰 Цена vs Стоимость")
            fig_price = go.Figure()
            fig_price.add_trace(go.Bar(
                x=df_offers["Name"],
                y=df_offers["Pack Price ($)"],
                name="Pack Price",
                marker_color='indianred'
            ))
            fig_price.add_trace(go.Bar(
                x=df_offers["Name"],
                y=df_offers["Base Value ($)"],
                name="Base Value",
                marker_color='lightsalmon'
            ))
            fig_price.update_layout(
                title="Цена пака vs Базовая стоимость",
                barmode='group'
            )
            st.plotly_chart(fig_price, use_container_width=True)

        st.subheader("📈 Скидки по офферам")
        fig_discount = px.line(
            df_offers,
            x="Name",
            y="Discount (%)",
            markers=True,
            title="Скидки по офферам"
        )
        st.plotly_chart(fig_discount, use_container_width=True)

        st.subheader("🎯 Типы паков")
        pack_types = df_offers["Pack Type"].value_counts()
        fig_types = px.pie(
            values=pack_types.values,
            names=pack_types.index,
            title="Распределение типов паков"
        )
        st.plotly_chart(fig_types, use_container_width=True)
    else:
        st.info("Нет офферов для сравнения. Создайте офферы в разделе 'Новый оффер'")

# ============== TAB 4: Настройки ==============
with tabs[3]:
    st.header("⚙️ Настройки базовых цен")

    st.info("Измените базовые цены за единицу ресурса. Это повлияет на все последующие расчёты.")

    settings_col1, settings_col2 = st.columns(2)

    with settings_col1:
        st.subheader("Ресурсы")
        new_gems = st.number_input("GEMS ($)", value=st.session_state.base_rates.gems, step=0.0001, format="%.4f")
        new_skip = st.number_input("SKIP ($)", value=st.session_state.base_rates.skip, step=0.01, format="%.2f")
        new_tnt = st.number_input("TNT ($)", value=st.session_state.base_rates.tnt, step=0.01, format="%.2f")
        new_nitro = st.number_input("NITRO ($)", value=st.session_state.base_rates.nitro, step=0.01, format="%.2f")
        new_no_ads = st.number_input("NO ADS ($)", value=st.session_state.base_rates.no_ads, step=0.1, format="%.1f")

    with settings_col2:
        st.subheader("Сундуки")
        new_small_chest = st.number_input("Small Chest ($)", value=st.session_state.base_rates.small_chest, step=0.01, format="%.2f")
        new_big_chest = st.number_input("Big Chest ($)", value=st.session_state.base_rates.big_chest, step=0.01, format="%.2f")
        new_no_chest = st.number_input("No Chest ($)", value=st.session_state.base_rates.no_chest, step=0.01, format="%.2f")

    if st.button("💾 Сохранить настройки", type="primary"):
        st.session_state.base_rates = BaseRates(
            gems=new_gems,
            skip=new_skip,
            tnt=new_tnt,
            nitro=new_nitro,
            no_ads=new_no_ads,
            small_chest=new_small_chest,
            big_chest=new_big_chest,
            no_chest=new_no_chest
        )
        st.success("✅ Настройки сохранены!")

    if st.button("🔄 Сбросить к значениям по умолчанию"):
        st.session_state.base_rates = BaseRates()
        st.success("✅ Настройки сброшены!")
        st.rerun()

# ============== TAB 5: Экспорт ==============
with tabs[4]:
    st.header("💾 Экспорт данных")

    if st.session_state.offers:
        st.subheader("Экспорт в Excel с формулами")

        excel_filename = st.text_input("Имя файла", value="offers_export.xlsx")

        if st.button("📥 Создать Excel файл", type="primary"):
            wb = Workbook()
            ws = wb.active
            ws.title = "Offers"

            # Базовые цены в первой строке (синий текст для входных данных)
            ws['A1'] = "За 1 одиницю ($) -"
            ws['B1'] = st.session_state.base_rates.gems
            ws['C1'] = st.session_state.base_rates.skip
            ws['D1'] = st.session_state.base_rates.tnt
            ws['E1'] = st.session_state.base_rates.nitro
            ws['F1'] = st.session_state.base_rates.no_ads

            for cell in ['B1', 'C1', 'D1', 'E1', 'F1']:
                ws[cell].font = Font(color="0000FF", bold=True)

            # Заголовки (жирный текст)
            headers = ["Name", "GEMS", "SKIP", "TNT", "NITRO", "NO ADS", "Chest Type", "Chest Amount",
                      "Pack Price ($)", "Base Value ($)", "Discount (%)", "ROI (%)",
                      "Value Badge", "Multiplier", "Min SkipIts", "Pack Type"]

            for idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Данные офферов с формулами
            for row_idx, offer in enumerate(st.session_state.offers, start=4):
                ws.cell(row=row_idx, column=1, value=offer["Name"])
                ws.cell(row=row_idx, column=2, value=offer["Gems"])
                ws.cell(row=row_idx, column=3, value=offer["Skip"])
                ws.cell(row=row_idx, column=4, value=offer["TNT"])
                ws.cell(row=row_idx, column=5, value=offer["Nitro"])
                ws.cell(row=row_idx, column=6, value=offer["No Ads"])
                ws.cell(row=row_idx, column=7, value=offer["Chest Type"])
                ws.cell(row=row_idx, column=8, value=offer["Chest Amount"])
                ws.cell(row=row_idx, column=9, value=offer["Pack Price ($)"])

                # Base Value - формула (чёрный текст)
                base_value_formula = f"=B{row_idx}*$B$1+C{row_idx}*$C$1+D{row_idx}*$D$1+E{row_idx}*$E$1+F{row_idx}*$F$1"
                ws.cell(row=row_idx, column=10, value=base_value_formula)

                # Discount - формула
                discount_formula = f"=IF(J{row_idx}=0,0,(J{row_idx}-I{row_idx})/J{row_idx}*100)"
                ws.cell(row=row_idx, column=11, value=discount_formula)

                # ROI - формула
                roi_formula = f"=IF(I{row_idx}=0,0,(J{row_idx}-I{row_idx})/I{row_idx}*100)"
                ws.cell(row=row_idx, column=12, value=roi_formula)

                ws.cell(row=row_idx, column=13, value=offer["Value Badge"])
                ws.cell(row=row_idx, column=14, value=offer["Multiplier"])
                ws.cell(row=row_idx, column=15, value=offer["Min SkipIts"])
                ws.cell(row=row_idx, column=16, value=offer["Pack Type"])

            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(excel_filename)

            st.success(f"✅ Excel файл создан: {excel_filename}")

            # Пересчёт формул
            if st.button("🔄 Пересчитать формулы (LibreOffice)"):
                result = subprocess.run(['python3', 'recalc.py', excel_filename],
                                      capture_output=True, text=True)
                try:
                    recalc_result = json.loads(result.stdout)
                    if recalc_result.get('status') == 'success':
                        st.success(f"✅ Формулы пересчитаны! Всего формул: {recalc_result.get('total_formulas', 0)}")
                    elif recalc_result.get('status') == 'errors_found':
                        st.error(f"❌ Найдены ошибки: {recalc_result.get('total_errors', 0)}")
                        st.json(recalc_result.get('error_summary', {}))
                    else:
                        st.error(f"❌ Ошибка: {recalc_result.get('error', 'Unknown error')}")
                except json.JSONDecodeError:
                    st.error(f"Ошибка парсинга результата: {result.stdout}")

            # Скачивание файла
            with open(excel_filename, "rb") as file:
                st.download_button(
                    label="⬇️ Скачать Excel файл",
                    data=file,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        st.divider()

        st.subheader("Экспорт в CSV")
        df_offers = pd.DataFrame(st.session_state.offers)
        csv = df_offers.to_csv(index=False).encode('utf-8')

        st.download_button(
            label="⬇️ Скачать CSV",
            data=csv,
            file_name="offers_export.csv",
            mime="text/csv"
        )
    else:
        st.info("Нет офферов для экспорта. Создайте офферы в разделе 'Новый оффер'")

st.divider()
st.markdown("---")
st.caption("💎 Offer Calculator v1.0 | Made with Streamlit")
